from func.base import month_name
import datetime
import pprint
from data.config import URL_GOOGLE_DOC_DEV, URL_GOOGLE_DOC_SUP
import requests
import openpyxl
import io
import time


def always():
    while True:
        print("'I'm slepping 10s ")
        time.sleep(10)

def get_color():
    now=datetime.datetime.now()
    response1 = requests.get(URL_GOOGLE_DOC_DEV,stream=True)
    google_doc_sup = openpyxl.load_workbook(filename=io.BytesIO(response1.content), data_only=True)
    list_name = [f"{month_name(now.month)} {now.year}"]
    print(list_name)
    sheet =google_doc_sup[f"{month_name(now.month)} {now.year}"]
    cell_obj = sheet.cell(row=58, column=4)
    color = cell_obj.fill.start_color.index
    value = cell_obj.value
    print(color)
    print(value)
    print(type(value))
    cell_obj_2 = sheet.cell(row=58, column=2)
    color_2 = cell_obj_2.fill.start_color.index
    value_2 = cell_obj_2.value
    print(color_2)
    print(value_2)
    print(type(value_2))


    # print(now.date())
    # print(now.date()==value.date())
if __name__ == '__main__':
    get_color()
    # always()
    # print("Next function")