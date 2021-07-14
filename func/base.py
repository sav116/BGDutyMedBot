import os

import openpyxl, io, requests, datetime
from data.config import URL_GOOGLE_DOC_DEV, URL_GOOGLE_DOC_SUP


def get_weekday(day):
    weekdays_dict = {
        0: 'Понедельник',
        1: 'Вторник',
        2: 'Среда',
        3: 'Четверг',
        4: 'Пятница',
        5: 'Суббота',
        6: 'Воскресенье'
    }
    return weekdays_dict[day]


def update_dev_dict(doc):
    """
        developers = {
            '@username':
                {
                    'phone': '790000000',
                    'fio': 'Ivanov Ivan Ivanovich',
                    'timezone': 'МСК+0'
                }
        }
        """
    now = datetime.datetime.now()
    developers = {}
    try:
        sheet_dev = doc[f"{month_name(now.month)} {now.year}"]
    except:
        return None
    for i in range(15, 140):
        cell_fio = sheet_dev.cell(row=i, column=2).value  # колонка с фио
        cell_phone = sheet_dev.cell(row=i, column=3).value
        cell_nick = sheet_dev.cell(row=i, column=4).value
        cell_timezone = sheet_dev.cell(row=i, column=6).value
        cell_direction = sheet_dev.cell(row=i, column=7).value

        if isinstance(cell_nick, str):
            if cell_nick.strip().startswith('@'):
                developers[cell_nick.strip()] = {
                    'phone': str(cell_phone).strip(),
                    'fio': str(cell_fio).strip(),
                    'timezone': str(cell_timezone).strip(),
                    'direction': str(cell_direction).strip()
                }
    return developers


def update_sup_dict(doc):
    """
        developers = {
            '@username':
                {
                    'phone': '790000000',
                    'fio': 'Ivanov Ivan Ivanovich',
                    'timezone': 'МСК+0'
                }
        }
        """
    now = datetime.datetime.now()
    technical_support = {}
    try:
        sheet_sup = doc[f"{month_name(now.month)} {now.year}"]
    except:
        return None
    for i in range(15, 40):
        cell_fio = sheet_sup.cell(row=i, column=2).value
        cell_phone = sheet_sup.cell(row=i, column=3).value
        cell_nick = sheet_sup.cell(row=i, column=4).value
        cell_timezone = sheet_sup.cell(row=i, column=6).value
        if isinstance(cell_nick, str):
            if cell_nick.strip().startswith('@'):
                technical_support[cell_nick.strip()] = {
                    'phone': str(cell_phone).strip(),
                    'fio': str(cell_fio).strip(),
                    'timezone': str(cell_timezone).strip()
                }
    return technical_support


def _download_docs():
    _GOOGLE_DEV = None
    _GOOGLE_SUP = None

    _DATA_DEV = None
    _DATA_SUP = None

    _GOOGLE_DEV_SHEET_CUR_MONTH = None
    _GOOGLE_DEV_SHEET_NEXT_MONTH = None

    _GOOGLE_SUP_SHEET_CUR_MONTH = None
    _GOOGLE_SUP_SHEET_NEXT_MONTH = None

    now = datetime.datetime.now()

    try:
        # Пытаемся открыть гугл док с разработчиками (3 линия) и обновить данные сотрудников
        response1 = requests.get(
            URL_GOOGLE_DOC_DEV,
            stream=True)
        _GOOGLE_DEV = openpyxl.load_workbook(filename=io.BytesIO(response1.content), data_only=True)
        _DATA_DEV = update_dev_dict(_GOOGLE_DEV)
    except:
        print("Не удалось открыть гугл док с разработчиками (3 линия) и обновить данные сотрудников")

    try:
        # Пытаемся открыть книгу с разработчиками, актуальную в данный момент
        _GOOGLE_DEV_SHEET_CUR_MONTH = _GOOGLE_DEV[f"{month_name(now.month)} {now.year}"]
    except:
        print("Не удалось открыть книгу с разработчиками, актуальную в данный момент")

    try:
        # Пытаемся открыть гугл док с техподдержкой (1 линия) и обновить данные сотрудников
        response2 = requests.get(
            URL_GOOGLE_DOC_SUP,
            stream=True)
        _GOOGLE_SUP = openpyxl.load_workbook(filename=io.BytesIO(response2.content), data_only=True)
        _DATA_SUP = update_sup_dict(_GOOGLE_SUP)
    except:
        print("Не удалось открыть гугл док с техподдержкой (1 линия) и обновить данные сотрудников")

    try:
        # Пытаемся открыть открыть книгу с техподдержкой актуальную в данный момент
        _GOOGLE_SUP_SHEET_CUR_MONTH = _GOOGLE_SUP[f"{month_name(now.month)} {now.year}"]
    except:
        print("Не удалось открыть открыть книгу с техподдержкой, актуальную в данный момент")

    try:
        # Пытаемся открыть книгу с разработчиками, на следующий месяц
        _GOOGLE_DEV_SHEET_NEXT_MONTH = _GOOGLE_DEV[f"{month_name(now.month + 1)} {now.year}"]
    except:
        print("Не удалось открыть книгу с разработчиками, на следующий месяц")

    try:
        # Пытаемся открыть открыть книгу с техподдержкой на следующий месяц
        _GOOGLE_SUP_SHEET_NEXT_MONTH = _GOOGLE_SUP[f"{month_name(now.month + 1)} {now.year}"]
    except:
        print("Не удалось открыть открыть книгу с техподдержкой на следующий месяц")

    return _GOOGLE_DEV, _GOOGLE_SUP, _DATA_DEV, _DATA_SUP, _GOOGLE_DEV_SHEET_CUR_MONTH, _GOOGLE_DEV_SHEET_NEXT_MONTH, _GOOGLE_SUP_SHEET_CUR_MONTH, _GOOGLE_SUP_SHEET_NEXT_MONTH


def month_name(month):
    months = ['Январь', 'Февраль', 'Март', 'Апрель', 'Май', 'Июнь', 'Июль', 'Август', 'Сентябрь', 'Октябрь', 'Ноябрь',
              'Декабрь']
    if (0 < month < 13):
        return months[month - 1]
    else:
        return "====bad month==="


def max_day_of_month(any_day):
    next_month = any_day.replace(day=28) + datetime.timedelta(days=4)  # this will never fail
    return next_month - datetime.timedelta(days=next_month.day)


def vladivostoc_time_sup():
    start_work_day = datetime.time(00, 00, 00)
    end_work_day = datetime.time(8, 00, 00)
    now = datetime.datetime.now()
    if now.weekday() < 5 and start_work_day <= now.time() <= end_work_day:
        return True
    return False


def working_day_sup():
    start_work_day = datetime.time(8, 00, 00)
    end_work_day = datetime.time(18, 00, 00)
    now = datetime.datetime.now()
    if now.weekday() < 5 and start_work_day <= now.time() <= end_work_day:
        return True
    return False


def working_day_dev():
    start_work_day = datetime.time(5, 00, 00)
    end_work_day = datetime.time(18, 00, 00)
    now = datetime.datetime.now()
    if now.weekday() < 5 and start_work_day <= now.time() <= end_work_day:
        return True
    return False


def not_working_time_weekdays():
    start_work_day = datetime.time(18, 00, 00)
    end_work_day = datetime.time(00, 00, 00)
    now = datetime.datetime.now()
    if now.weekday() < 5 and start_work_day <= now.time() <= end_work_day:
        return True
    return False


def day_off():
    now = datetime.datetime.now()
    if now.weekday() > 4:
        return True
    return False


def day_off_dev():
    from loader import GOOGLE_DEV_SHEET_CUR_MONTH
    now = datetime.datetime.now()
    if now.weekday() > 4 or GOOGLE_DEV_SHEET_CUR_MONTH.cell(row=1, column=now.day + 3) == "FFFFE599":
        return True
    return False


def get_nic_telegramm(surname, sheet):
    for row in range(18, 45):
        if isinstance(sheet.cell(row=row, column=2).value, str):
            cell_nick_value = sheet.cell(row=row, column=4).value
            cell_fio_value = sheet.cell(row=row, column=2).value.strip()
            if surname in cell_fio_value:
                return cell_nick_value
    return None


def get_nic_telegramm_dev(surname, sheet):
    for row in range(40, 130):
        cell_2_value = sheet.cell(row=row, column=2).value
        if isinstance(cell_2_value, str):
            cell_nick_value = sheet.cell(row=row, column=4).value
            cell_fio_value = sheet.cell(row=row, column=2).value.strip()
            if surname in cell_fio_value:
                return cell_nick_value
    return None