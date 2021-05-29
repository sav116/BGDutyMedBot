import openpyxl
import time
import requests
import io
import datetime

from data.config import URL_GOOGLE_DOC_DEV, URL_GOOGLE_DOC_SUP
from func.base import month_name

google_doc_dev = None
google_doc_sup = None

developers = {}
technical_support = {}


def update_dev_dict(doc):
    """
        developers = {
            '@username':
                {
                    'phone': '790000000',
                    'fio': 'Ivanov Ivan Ivanovich',
                    'timezone': 'МСК+0'
                }
        }
        """
    now=datetime.datetime.now()
    developers={}
    try:
        sheet_dev = doc[f"{month_name(now.month)} {now.year}"]
    except:
        return "Can't open book's sheet"
    for i in range(15, 140):
        cell_fio=sheet_dev.cell(row=i, column=2).value
        cell_phone=sheet_dev.cell(row=i, column=3).value
        cell_nick=sheet_dev.cell(row=i, column=4).value
        cell_timezone=sheet_dev.cell(row=i, column=6).value
        if isinstance(cell_nick, str):
            if cell_nick.strip().startswith('@'):
                developers[cell_nick.strip()]={
                    'phone': str(cell_phone).strip(),
                    'fio': str(cell_fio).strip(),
                    'timezone': str(cell_timezone).strip()
                }
    return developers

def update_sup_dict(doc):
    """
        developers = {
            '@username':
                {
                    'phone': '790000000',
                    'fio': 'Ivanov Ivan Ivanovich',
                    'timezone': 'МСК+0'
                }
        }
        """
    now = datetime.datetime.now()
    technical_support={}
    try:
        sheet_sup = doc[f"{month_name(now.month)} {now.year}"]
    except:
        return "Can't open book's sheet"
    for i in range(15, 40):
        cell_fio=sheet_sup.cell(row=i, column=2).value
        cell_phone=sheet_sup.cell(row=i, column=3).value
        cell_nick=sheet_sup.cell(row=i, column=4).value
        cell_timezone=sheet_sup.cell(row=i, column=6).value
        if isinstance(cell_nick, str):
            if cell_nick.strip().startswith('@'):
                technical_support[cell_nick.strip()]={
                    'phone': str(cell_phone).strip(),
                    'fio': str(cell_fio).strip(),
                    'timezone': str(cell_timezone).strip()
                }
    return technical_support

def update_docs():
    while True:
        global google_doc_dev, google_doc_sup, developers, technical_support
        response1=requests.get(
            URL_GOOGLE_DOC_DEV,
            stream=True)
        google_doc_dev=openpyxl.load_workbook(filename=io.BytesIO(response1.content), data_only=True)
        developers = update_dev_dict(google_doc_dev)
        print(developers)
        response2=requests.get(
            URL_GOOGLE_DOC_SUP,
            stream=True)
        google_doc_sup=openpyxl.load_workbook(filename=io.BytesIO(response2.content), data_only=True)
        technical_support = update_sup_dict(google_doc_sup)
        print(technical_support)
        print('Google таблицы загружены в оперативную память')
        time.sleep(300)

def forced_update():
    global google_doc_dev, google_doc_sup
    response1 = requests.get(
        URL_GOOGLE_DOC_DEV,
        stream=True)
    google_doc_dev = openpyxl.load_workbook(filename=io.BytesIO(response1.content), data_only=True)

    response2 = requests.get(
        URL_GOOGLE_DOC_SUP,
        stream=True)
    google_doc_sup = openpyxl.load_workbook(filename=io.BytesIO(response2.content), data_only=True)
    print('Google таблицы обновлены принудительно')